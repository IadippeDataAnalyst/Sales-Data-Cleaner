"""
Sales data cleaning pipeline + monthly report with visualization.
Usage: python clean_and_report.py <csv_file>
"""
import pandas as pd
import matplotlib.pyplot as plt
import sys
from pathlib import Path

def load_data(path):
    """Loads the CSV and shows basic info."""
    df = pd.read_csv(path)
    print(f"\n[1/5] Data loaded: {len(df)} rows, {len(df.columns)} columns")
    return df

def clean_data(df):
    """Removes duplicates, normalizes categories, parses dates, handles nulls."""
    initial_rows = len(df)
    
    # Remove duplicates
    df = df.drop_duplicates().reset_index(drop=True)
    print(f"[2/5] Duplicates removed: {initial_rows - len(df)}")
    
    # Normalize categories (strip spaces, capitalize)
    df['category'] = df['category'].str.strip().str.capitalize()
    
    # Parse mixed-format dates
    df['date'] = pd.to_datetime(df['date'], format='mixed', dayfirst=False, errors='coerce')
    invalid_dates = df['date'].isna().sum()
    if invalid_dates > 0:
        print(f"    Warning: {invalid_dates} invalid dates dropped")
        df = df.dropna(subset=['date']).reset_index(drop=True)
    
    # Fill nulls in quantity and price with median
    df['quantity'] = df['quantity'].fillna(df['quantity'].median())
    df['unit_price'] = df['unit_price'].fillna(df['unit_price'].median())
    
    # Calculate total revenue column
    df['revenue'] = df['quantity'] * df['unit_price']
    
    # Add month column for grouping
    df['month'] = df['date'].dt.to_period('M').astype(str)
    
    print(f"[3/5] Cleaning complete: {len(df)} final rows")
    return df

def generate_report(df, output_folder='report'):
    """Groups by month and category, generates chart and Excel with professional formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    
    Path(output_folder).mkdir(exist_ok=True)
    
    # Convert date to date-only (no time) before export
    df_export = df.copy()
    df_export['date'] = df_export['date'].dt.date
    
    # Pivot: month × category
    pivot = df.pivot_table(
        index='month',
        columns='category',
        values='revenue',
        aggfunc='sum',
        fill_value=0
    )
    
    # --- Chart ---
    fig, ax = plt.subplots(figsize=(12, 6))
    pivot.plot(kind='bar', stacked=True, ax=ax, colormap='tab20')
    ax.set_title('Monthly Revenue by Category', fontsize=14, fontweight='bold')
    ax.set_xlabel('Month')
    ax.set_ylabel('Revenue ($)')
    ax.legend(title='Category', bbox_to_anchor=(1.02, 1), loc='upper left')
    plt.xticks(rotation=45)
    plt.tight_layout()
    
    chart_path = f'{output_folder}/monthly_revenue.png'
    plt.savefig(chart_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"[4/5] Chart saved: {chart_path}")
    
    # --- Excel with professional formatting ---
    excel_path = f'{output_folder}/sales_report.xlsx'
    top_products = df.groupby('product')['revenue'].sum().sort_values(ascending=False).head(10)
    
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, sheet_name='Clean Data', index=False)
        pivot.to_excel(writer, sheet_name='Monthly Summary')
        top_products.to_frame('Total Revenue').to_excel(writer, sheet_name='Top 10 Products')
        
        workbook = writer.book
        
        # Header style
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_fill = PatternFill(start_color='305496', end_color='305496', fill_type='solid')
        header_align = Alignment(horizontal='center', vertical='center')
        
        money_format = '"$"#,##0.00'
        date_format = 'yyyy-mm-dd'
        
        money_columns = {
            'Clean Data': ['unit_price', 'revenue'],
            'Monthly Summary': 'all_except_first',
            'Top 10 Products': ['Total Revenue']
        }
        
        for sheet_name in ['Clean Data', 'Monthly Summary', 'Top 10 Products']:
            ws = workbook[sheet_name]
            
            # Format header row
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_align
            
            # Map column names to indices
            headers = {cell.value: cell.column for cell in ws[1]}
            
            # Apply money format
            if sheet_name == 'Monthly Summary':
                for col_idx in range(2, ws.max_column + 1):
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        for cell in row:
                            cell.number_format = money_format
            else:
                for col_name in money_columns[sheet_name]:
                    if col_name in headers:
                        col_idx = headers[col_name]
                        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                            for cell in row:
                                cell.number_format = money_format
            
            # Date format in "Clean Data"
            if sheet_name == 'Clean Data' and 'date' in headers:
                col_idx = headers['date']
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        cell.number_format = date_format
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                ws.column_dimensions[column_letter].width = max_length + 3
            
            # Freeze header row
            ws.freeze_panes = 'A2'
    
    print(f"[5/5] Excel saved: {excel_path}")
    
    # Summary metrics in console
    print(f"\n{'='*50}")
    print("EXECUTIVE SUMMARY")
    print(f"{'='*50}")
    print(f"Total revenue: ${df['revenue'].sum():,.2f}")
    print(f"Average ticket: ${df['revenue'].mean():,.2f}")
    print(f"Period: {df['date'].min().date()} to {df['date'].max().date()}")
    print(f"Leading category: {df.groupby('category')['revenue'].sum().idxmax()}")

def main():
    file = sys.argv[1] if len(sys.argv) > 1 else 'messy_sales.csv'
    df = load_data(file)
    clean_df = clean_data(df)
    generate_report(clean_df)

if __name__ == '__main__':
    main()
